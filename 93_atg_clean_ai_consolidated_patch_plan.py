#!/usr/bin/env python3
"""Build a read-only consolidated ATG clean-AI-space patch plan.

No Geo writes. No ops are created.

The goal is to answer:
- which expected batch A-D agent cards already exist in clean AI space;
- which canonical card should be enriched when duplicates exist;
- which agent fields/relations/blocks are missing;
- which approved local cover/avatar assets exist and how they differ from live;
- which taxonomy visuals are still missing from live.

Run:
  python3 93_atg_clean_ai_consolidated_patch_plan.py
"""

from __future__ import annotations

import json
import os
import sys
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - the repo already uses openpyxl.
    load_workbook = None


ROOT = Path(__file__).resolve().parent
GEO_ATG = ROOT / "geo_atg"
VISUALS = GEO_ATG / "visuals"
SPACE_ID = os.environ.get("SPACE_ID", "41e851610e13a19441c4d980f2f2ce6b")
API_URL = os.environ.get("GEO_GQL_URL", "https://testnet-api.geobrowser.io/graphql")
DATE = "2026-06-26"

OUT_JSON = GEO_ATG / f"atg_clean_ai_consolidated_patch_plan_{DATE}.json"
OUT_MD = GEO_ATG / f"atg_clean_ai_consolidated_patch_plan_{DATE}.md"


REL = {
    "Types": "8f151ba4de204e3c9cb499ddf96f48f1",
    "Capabilities": "15f630ff5d87f0e3992fdd68567970fc",
    "Cover": "34f535072e6b42c5a84443981a77cfa2",
    "Developers": "b36bba262b6b45dabe8b6fe1d41f5f96",
    "Features": "b3c9e2e050234801a0da05314326c439",
    "Avatar": "1155befffad549b7a2e0da4777b8792c",
    "LicenseStatus": "a79b63d453dd4938a3091ff448d36cbc",
    "Models": "968742a40c109c633de1c70a79587487",
    "Protocols": "27f4e74e301943a89dc573bbb6df8f00",
    "ProvidedBy": "9e6512b649d0daa76d9d6b0acd3ffdcc",
    "Repositories": "a25e04e2656f4700ba10e31b231b446a",
    "SoftwareLicenses": "7fc423a18e304205988af467d1f8b84a",
    "UseCases": "a5d4d84003dd43e780d300770141388d",
    "Blocks": "beaba5cba67741a8b35377030613fc70",
}

PROP = {
    "GitHub": "9eedefa860ae4ac19a04805054a4b094",
    "Docs": "a446528df6b24ecab04bc4dd7dedfbd9",
    "X": "0d6259784b3c4b57a86fde45c997c73c",
    "GitHubStars": "a79523e6aaa5dfc0701711234df6af9d",
    "ReleaseDate": "1170c5a68a015fb2addc44382fe7d0f5",
    "ActivelyMaintained": "ffab986cc06f6f288efae9db7c4a3fdc",
    "SkillsCount": "45509a98a0a2473ba727f7b170ed813b",
    "Website": "eed38e74e67946bf8a42ea3e4f8fb5fb",
}

AGENT_TYPE_ID = "9069cd7680cabc7b5e7aace5bc0da4d3"
PROJECT_TYPE_ID = "484a18c5030a499cb0f2ef588ff16d50"
TOOL_TYPE_ID = "fa464fe0c27b4d54bbac4caa20ca7781"

APPROVED_COVER_SOURCES = {
    "quality_gate_fixed_cover",
    "brand_preserving_cover",
    "small_brand_icon_cover",
    "manual_external_candidate",
    "final_user_url_cover",
    "finalized_raw_cover",
}
APPROVED_AVATAR_SOURCES = {
    "github",
    "site_icon",
    "site_og",
    "quality_gate_fixed_avatar",
    "generated_neutral_atg",
}


def read_json(path: Path, fallback: Any = None) -> Any:
    if not path.exists():
        return fallback
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def chunks(values: list[Any], size: int) -> list[list[Any]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def gql(query: str, variables: dict[str, Any] | None = None) -> dict[str, Any]:
    body = json.dumps({"query": query, "variables": variables or {}}).encode()
    request = urllib.request.Request(
        API_URL,
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        payload = json.loads(response.read().decode())
    if payload.get("errors"):
        raise RuntimeError(json.dumps(payload["errors"], indent=2))
    return payload["data"]


def entity_fragment() -> str:
    return """
      id
      name
      description
      createdAt
      updatedAt
      values(first: 1000) {
        nodes {
          propertyId
          text
          integer
          decimal
          float
          boolean
          date
          datetime
          time
        }
      }
      relations(first: 1000) {
        nodes {
          id
          typeId
          toEntity { id name }
        }
      }
    """


def query_entities_by_names(names: list[str]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for part in chunks(sorted(set(name for name in names if name)), 45):
        data = gql(
            f"""
            query($space: UUID!, $names: [String!]!) {{
              entities(spaceId: $space first: 1000 filter: {{ name: {{ in: $names }} }}) {{
                {entity_fragment()}
              }}
            }}
            """,
            {"space": SPACE_ID, "names": part},
        )
        result.extend(data.get("entities") or [])
    return result


def query_entities_by_ids(ids: list[str]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for part in chunks(sorted(set(entity_id for entity_id in ids if entity_id)), 45):
        data = gql(
            f"""
            query($space: UUID!, $ids: [UUID!]!) {{
              entities(spaceId: $space first: 1000 filter: {{ id: {{ in: $ids }} }}) {{
                {entity_fragment()}
              }}
            }}
            """,
            {"space": SPACE_ID, "ids": part},
        )
        result.extend(data.get("entities") or [])
    return result


def rels(entity: dict[str, Any], rel_type: str) -> list[dict[str, Any]]:
    return [
        {
            "relation_id": item["id"],
            "to_id": item["toEntity"]["id"],
            "to_name": item["toEntity"].get("name"),
        }
        for item in entity.get("relations", {}).get("nodes", [])
        if item.get("typeId") == rel_type
    ]


def has_rel(entity: dict[str, Any], rel_type: str, to_id: str) -> bool:
    return any(item["to_id"] == to_id for item in rels(entity, rel_type))


def value_raw(item: dict[str, Any]) -> Any:
    for key in ("text", "integer", "decimal", "float", "boolean", "date", "datetime", "time"):
        if item.get(key) is not None:
            return item[key]
    return None


def values(entity: dict[str, Any], prop_id: str) -> list[Any]:
    out: list[Any] = []
    for item in entity.get("values", {}).get("nodes", []):
        if item.get("propertyId") != prop_id:
            continue
        raw = value_raw(item)
        if raw is not None and str(raw).strip() != "":
            out.append(raw)
    return out


def block_names(entity: dict[str, Any]) -> set[str]:
    return {item["to_name"] for item in rels(entity, REL["Blocks"]) if item.get("to_name")}


def created_at(entity: dict[str, Any]) -> int:
    try:
        return int(entity.get("createdAt") or 0)
    except (TypeError, ValueError):
        return 0


def load_batch_agents() -> tuple[list[dict[str, Any]], dict[str, str]]:
    agents: list[dict[str, Any]] = []
    batch_by_id: dict[str, str] = {}
    for batch in ("A", "B", "C", "D"):
        path = GEO_ATG / f"atg_agents_batch{batch}_publish.json"
        for row in read_json(path, []):
            row = dict(row)
            row["batch"] = batch
            agents.append(row)
            batch_by_id[row["id"]] = batch
    return agents, batch_by_id


def load_workbook_edges(agent_ids: set[str]) -> dict[str, dict[str, list[str]]]:
    edge_counts: dict[str, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    workbook = GEO_ATG / "source_workbooks" / "agents_v13_publish_layers.xlsx"
    if load_workbook is None or not workbook.exists():
        return edge_counts
    wb = load_workbook(workbook, read_only=True, data_only=True)
    sheets = [
        "Edges_Benchmarks",
        "Edges_Runtimes",
        "Edges_MemoryTypes",
        "Edges_RiskSurface",
        "Edges_RiskMitigations",
        "Edges_Domains",
    ]
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=0)}
        src_idx = headers.get("source_id")
        target_idx = headers.get("target_id")
        if src_idx is None or target_idx is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            source_id = row[src_idx]
            target_id = row[target_idx]
            if source_id in agent_ids and target_id:
                edge_counts[str(source_id)][sheet_name].append(str(target_id))
    return edge_counts


def load_workbook_specs() -> dict[str, dict[str, dict[str, Any]]]:
    specs: dict[str, dict[str, dict[str, Any]]] = {}
    workbook = GEO_ATG / "source_workbooks" / "agents_v13_publish_layers.xlsx"
    if load_workbook is None or not workbook.exists():
        return specs
    wb = load_workbook(workbook, read_only=True, data_only=True)
    for sheet_name in [
        "Capabilities",
        "Protocols",
        "Models",
        "Licenses",
        "Domains",
        "Runtimes",
        "MemoryTypes",
        "RiskTypes",
        "Benchmarks",
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration:
            continue
        sheet_specs: dict[str, dict[str, Any]] = {}
        for row in rows:
            item = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row))) if headers[idx]}
            raw_id = item.get("id")
            if raw_id is None:
                continue
            slug = str(raw_id).strip()
            if not slug or slug == "slug":
                continue
            name = item.get("name")
            if name is None or not str(name).strip():
                continue
            spec: dict[str, Any] = {
                "id": slug,
                "name": str(name).strip(),
                "description": str(item.get("description") or "").strip(),
            }
            for key in ("source_url", "spec_url", "leaderboard_url", "url"):
                if item.get(key):
                    spec[key] = str(item[key]).strip()
            sheet_specs[slug] = spec
        specs[sheet_name] = sheet_specs
    return specs


def visual_key_from_asset_id(asset_id: str, suffix: str) -> str:
    if asset_id.endswith(suffix):
        return asset_id[: -len(suffix)]
    return asset_id


def load_visual_plan(agents: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    by_id: dict[str, dict[str, Any]] = {agent["id"]: {"agent_id": agent["id"], "name": agent["name"]} for agent in agents}
    name_to_id = {agent["name"]: agent["id"] for agent in agents}

    def merge_record(record: dict[str, Any]) -> None:
        agent_id = record.get("agent_id")
        if not agent_id:
            return
        slot = by_id.setdefault(agent_id, {"agent_id": agent_id, "name": record.get("display_name")})
        for key in (
            "cover_file",
            "avatar_file",
            "cover_source_type",
            "avatar_source_type",
            "cover_review_status",
            "avatar_review_status",
            "visual_review_status",
            "approved",
            "status",
        ):
            if record.get(key) is not None:
                slot[key] = record.get(key)
        if record.get("cover_path") is not None:
            slot["cover_file"] = record.get("cover_path")
        if record.get("avatar_path") is not None:
            slot["avatar_file"] = record.get("avatar_path")

    for path in [
        VISUALS / "atg_batchA_prod_visual_manifest_2026-06-24.json",
        VISUALS / "atg_showcase37_visual_manifest_2026-06-23.json",
        VISUALS / "atg_remaining_visual_attach_manifest_2026-06-21.json",
    ]:
        data = read_json(path, {})
        for record in data.get("records", []):
            merge_record(record)

    for batch in ("B", "C", "D"):
        cover_data = read_json(VISUALS / f"atg_batch{batch}_cover_attach_manifest_2026-06-23.json", {})
        if batch in ("C", "D"):
            cover_data = read_json(VISUALS / f"atg_batch{batch}_cover_attach_manifest_2026-06-24.json", {})
        for asset in cover_data.get("assets", []):
            agent_id = name_to_id.get(asset.get("target_entity_name"))
            if not agent_id:
                agent_id = visual_key_from_asset_id(asset.get("id", ""), "-cover")
            slot = by_id.setdefault(agent_id, {"agent_id": agent_id, "name": asset.get("target_entity_name")})
            slot["cover_file"] = asset.get("file")
            slot.setdefault("cover_source_type", "split_cover_manifest")
            slot["cover_manifest"] = str(cover_data.get("version", ""))

        avatar_data = read_json(VISUALS / f"atg_batch{batch}_avatar_attach_manifest_2026-06-23.json", {})
        if batch in ("C", "D"):
            avatar_data = read_json(VISUALS / f"atg_batch{batch}_avatar_attach_manifest_2026-06-24.json", {})
        for asset in avatar_data.get("assets", []):
            agent_id = name_to_id.get(asset.get("target_entity_name"))
            if not agent_id:
                agent_id = visual_key_from_asset_id(asset.get("id", ""), "-avatar")
            slot = by_id.setdefault(agent_id, {"agent_id": agent_id, "name": asset.get("target_entity_name")})
            slot["avatar_file"] = asset.get("file")
            slot.setdefault("avatar_source_type", "split_avatar_manifest")
            slot["avatar_manifest"] = str(avatar_data.get("version", ""))

    return by_id


def visual_is_approved(item: dict[str, Any]) -> bool:
    if item.get("approved") is True:
        return True
    if item.get("visual_review_status") == "approved":
        return True
    if item.get("cover_review_status") == "approved" and item.get("avatar_review_status") == "approved":
        return True
    cover_source = item.get("cover_source_type")
    avatar_source = item.get("avatar_source_type")
    if cover_source in {"split_cover_manifest"}:
        # Split batch manifests were generated from the curated visual pipeline.
        cover_ok = True
    else:
        cover_ok = cover_source in APPROVED_COVER_SOURCES
    if avatar_source in {"split_avatar_manifest"}:
        avatar_ok = True
    else:
        avatar_ok = avatar_source in APPROVED_AVATAR_SOURCES
    return bool(cover_ok and avatar_ok)


def file_state(path_value: str | None) -> dict[str, Any]:
    if not path_value:
        return {"path": None, "exists": False}
    path = ROOT / path_value
    return {"path": path_value, "exists": path.exists(), "size": path.stat().st_size if path.exists() else None}


def scalar_expected(agent: dict[str, Any]) -> dict[str, Any]:
    return {
        "Website": agent.get("website_url") or "",
        "GitHub": agent.get("github_url") or "",
        "Docs": agent.get("docs_url") or "",
        "X": agent.get("x_url") or "",
        "GitHubStars": agent.get("stars_at_collection"),
        "ReleaseDate": agent.get("release_date") or "",
        "ActivelyMaintained": True,
        "SkillsCount": agent.get("skills_count"),
    }


def normalize_scalar(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def planned_scalar_updates(agent: dict[str, Any], entity: dict[str, Any] | None) -> list[dict[str, Any]]:
    updates: list[dict[str, Any]] = []
    expected = scalar_expected(agent)
    if entity is None:
        return [{"field": key, "action": "set_on_create", "value": value} for key, value in expected.items() if normalize_scalar(value)]
    for field, expected_value in expected.items():
        prop_id = PROP[field]
        current_values = values(entity, prop_id)
        current = current_values[0] if current_values else ""
        if normalize_scalar(expected_value) and normalize_scalar(current) != normalize_scalar(expected_value):
            updates.append({"field": field, "current": current, "desired": expected_value})
        if not normalize_scalar(expected_value) and current_values and field in {"Website", "GitHub", "Docs", "X"}:
            updates.append({"field": field, "current": current, "desired": "", "action": "unset_null_or_empty_source"})
    return updates


def get_type_ids(agent: dict[str, Any]) -> list[str]:
    # The clean AI space agent-card pattern is Agent + Project + Tool.
    # Hermes Agent also has Product, but Product is not currently the broad ATG default.
    return [PROJECT_TYPE_ID, AGENT_TYPE_ID, TOOL_TYPE_ID]


def relation_plan(agent: dict[str, Any], entity: dict[str, Any] | None, edge_counts: dict[str, list[str]]) -> dict[str, Any]:
    if entity is None:
        return {
            "create_with": {
                "types": len(get_type_ids(agent)),
                "capabilities": len(agent.get("capabilities") or []),
                "features": len(agent.get("features") or []),
                "protocols": len(agent.get("protocols") or []),
                "models": len(agent.get("models") or []),
                "repository": bool(agent.get("github_url")),
                "use_cases": len(agent.get("domains") or []) or bool(agent.get("primary_use_case")),
            }
        }
    missing_types = [type_id for type_id in get_type_ids(agent) if not has_rel(entity, REL["Types"], type_id)]
    block_set = block_names(entity)
    return {
        "missing_type_relations": missing_types,
        "needs_developers": len(rels(entity, REL["Developers"])) == 0,
        "needs_provided_by": len(rels(entity, REL["ProvidedBy"])) == 0,
        "needs_repository": bool(agent.get("github_url")) and len(rels(entity, REL["Repositories"])) == 0,
        "needs_use_cases": (bool(agent.get("primary_use_case")) or bool(agent.get("domains"))) and len(rels(entity, REL["UseCases"])) == 0,
        "capabilities_live": len(rels(entity, REL["Capabilities"])),
        "capabilities_expected": len(agent.get("capabilities") or []),
        "features_live": len(rels(entity, REL["Features"])),
        "features_expected": len(agent.get("features") or []),
        "protocols_live": len(rels(entity, REL["Protocols"])),
        "protocols_expected": len(agent.get("protocols") or []),
        "models_live": len(rels(entity, REL["Models"])),
        "models_expected": len(agent.get("models") or []),
        "license_live": len(rels(entity, REL["SoftwareLicenses"])),
        "license_expected": bool(agent.get("license") and agent.get("license") != "NOASSERTION"),
        "blocks_missing": [
            block_name
            for block_name, targets in {
                "Runtimes": edge_counts.get("Edges_Runtimes", []),
                "Memory types": edge_counts.get("Edges_MemoryTypes", []),
                "Risks": list(edge_counts.get("Edges_RiskSurface", [])) + list(edge_counts.get("Edges_RiskMitigations", [])),
                "Benchmarks": edge_counts.get("Edges_Benchmarks", []),
            }.items()
            if targets and block_name not in block_set
        ],
    }


def taxonomy_visual_files(kind: str, slug: str) -> tuple[Path, Path]:
    dasherized = slug.replace("_", "-")
    base = VISUALS / "generated" / "taxonomy_bootstrap"
    return base / f"{kind}-{dasherized}-cover.png", base / f"{kind}-{dasherized}-avatar.png"


def main() -> None:
    agents, batch_by_id = load_batch_agents()
    expected_names = [agent["name"] for agent in agents]
    expected_ids = {agent["id"] for agent in agents}
    edge_counts = load_workbook_edges(expected_ids)
    target_specs = load_workbook_specs()
    visual_plan = load_visual_plan(agents)

    live_entities = query_entities_by_names(expected_names)
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entity in live_entities:
        by_name[entity["name"]].append(entity)

    rows: list[dict[str, Any]] = []
    duplicate_cleanup: list[dict[str, Any]] = []
    visual_issues: list[dict[str, Any]] = []
    missing_create_count = 0
    enrich_existing_count = 0

    for agent in agents:
        candidates = by_name.get(agent["name"], [])
        agent_candidates = [entity for entity in candidates if has_rel(entity, REL["Types"], AGENT_TYPE_ID)]
        canonical = min(agent_candidates, key=created_at) if agent_candidates else None
        action = "enrich_existing" if canonical else "create_agent"
        if canonical:
            enrich_existing_count += 1
        else:
            missing_create_count += 1

        duplicate_agent_candidates = sorted(agent_candidates, key=created_at)
        if len(duplicate_agent_candidates) > 1:
            duplicate_cleanup.append(
                {
                    "name": agent["name"],
                    "canonical": duplicate_agent_candidates[0]["id"],
                    "cleanup_candidates": [entity["id"] for entity in duplicate_agent_candidates[1:]],
                    "createdAt": {entity["id"]: entity.get("createdAt") for entity in duplicate_agent_candidates},
                }
            )

        visual = visual_plan.get(agent["id"], {"agent_id": agent["id"], "name": agent["name"]})
        cover_state = file_state(visual.get("cover_file"))
        avatar_state = file_state(visual.get("avatar_file"))
        approved = visual_is_approved(visual)
        live_cover_count = len(rels(canonical, REL["Cover"])) if canonical else 0
        live_avatar_count = len(rels(canonical, REL["Avatar"])) if canonical else 0
        if not approved or not cover_state["exists"] or not avatar_state["exists"]:
            visual_issues.append(
                {
                    "agent_id": agent["id"],
                    "name": agent["name"],
                    "approved": approved,
                    "cover": cover_state,
                    "avatar": avatar_state,
                    "cover_source_type": visual.get("cover_source_type"),
                    "avatar_source_type": visual.get("avatar_source_type"),
                }
            )

        row_edge_counts = edge_counts.get(agent["id"], {})
        edge_sheet_to_spec_sheet = {
            "Edges_Runtimes": "Runtimes",
            "Edges_MemoryTypes": "MemoryTypes",
            "Edges_RiskSurface": "RiskTypes",
            "Edges_RiskMitigations": "RiskTypes",
            "Edges_Benchmarks": "Benchmarks",
            "Edges_Domains": "Domains",
        }
        rows.append(
            {
                "agent_id": agent["id"],
                "batch": agent["batch"],
                "name": agent["name"],
                "action": action,
                "canonical_entity_id": canonical["id"] if canonical else None,
                "candidate_count": len(candidates),
                "agent_candidate_count": len(agent_candidates),
                "canonical_createdAt": canonical.get("createdAt") if canonical else None,
                "scalar_updates": planned_scalar_updates(agent, canonical),
                "relation_plan": relation_plan(agent, canonical, row_edge_counts),
                "visual": {
                    "approved": approved,
                    "cover": cover_state,
                    "avatar": avatar_state,
                    "live_cover_count": live_cover_count,
                    "live_avatar_count": live_avatar_count,
                    "cover_source_type": visual.get("cover_source_type"),
                    "avatar_source_type": visual.get("avatar_source_type"),
                },
                "edge_counts": {sheet: len(targets) for sheet, targets in row_edge_counts.items()},
                "edge_targets": {
                    sheet: [
                        target_specs.get(edge_sheet_to_spec_sheet.get(sheet, ""), {}).get(
                            target_id,
                            {"id": target_id, "name": target_id, "description": ""},
                        )
                        for target_id in targets
                    ]
                    for sheet, targets in row_edge_counts.items()
                },
            }
        )

    capability_slugs = sorted({slug for agent in agents for slug in (agent.get("capabilities") or []) if slug != "mcp_client"})
    feature_slugs = sorted({slug for agent in agents for slug in (agent.get("features") or [])})
    taxonomy_visuals = []
    for kind, slugs in (("capability", capability_slugs), ("feature", feature_slugs)):
        for slug in slugs:
            cover, avatar = taxonomy_visual_files(kind, slug)
            taxonomy_visuals.append(
                {
                    "kind": kind,
                    "slug": slug,
                    "cover": str(cover.relative_to(ROOT)),
                    "cover_exists": cover.exists(),
                    "avatar": str(avatar.relative_to(ROOT)),
                    "avatar_exists": avatar.exists(),
                }
            )
    missing_taxonomy_visual_files = [item for item in taxonomy_visuals if not item["cover_exists"] or not item["avatar_exists"]]
    unique_existing_rows_by_entity: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row["action"] != "enrich_existing" or not row.get("canonical_entity_id"):
            continue
        unique_existing_rows_by_entity.setdefault(row["canonical_entity_id"], row)
    unique_existing_rows = list(unique_existing_rows_by_entity.values())
    existing_only_summary = {
        "batch_rows": len([row for row in rows if row["action"] == "enrich_existing"]),
        "unique_canonical_cards": len(unique_existing_rows),
        "create_rows_excluded_from_next_proposal": len([row for row in rows if row["action"] == "create_agent"]),
        "rows_needing_repositories": sum(1 for row in unique_existing_rows if row["relation_plan"].get("needs_repository")),
        "rows_needing_use_cases": sum(1 for row in unique_existing_rows if row["relation_plan"].get("needs_use_cases")),
        "rows_needing_provided_by": sum(1 for row in unique_existing_rows if row["relation_plan"].get("needs_provided_by")),
        "rows_needing_blocks": sum(1 for row in unique_existing_rows if row["relation_plan"].get("blocks_missing")),
        "rows_needing_type_normalization": sum(1 for row in unique_existing_rows if row["relation_plan"].get("missing_type_relations")),
        "local_visual_issues": len(
            [
                row
                for row in unique_existing_rows
                if row["visual"].get("approved")
                and (not row["visual"]["cover"]["exists"] or not row["visual"]["avatar"]["exists"])
            ]
        ),
    }
    approved_visual_rows = [row for row in rows if row["visual"].get("approved")]
    local_visual_complete_rows = [
        row
        for row in approved_visual_rows
        if row["visual"]["cover"]["exists"] and row["visual"]["avatar"]["exists"]
    ]
    existing_visual_rows = [row for row in approved_visual_rows if row["action"] == "enrich_existing"]
    create_visual_rows = [row for row in approved_visual_rows if row["action"] == "create_agent"]
    existing_missing_live_visual = [
        row
        for row in existing_visual_rows
        if row["visual"].get("live_cover_count", 0) == 0 or row["visual"].get("live_avatar_count", 0) == 0
    ]
    visual_diff_summary = {
        "approved_rows": len(approved_visual_rows),
        "local_visual_complete": len(local_visual_complete_rows),
        "local_visual_missing": len(approved_visual_rows) - len(local_visual_complete_rows),
        "existing_cards_with_approved_visuals": len(existing_visual_rows),
        "existing_cards_missing_live_cover_or_avatar": len(existing_missing_live_visual),
        "create_cards_with_approved_complete_visuals": sum(
            1
            for row in create_visual_rows
            if row["visual"]["cover"]["exists"] and row["visual"]["avatar"]["exists"]
        ),
        "cover_source_types": dict(Counter(row["visual"].get("cover_source_type") for row in approved_visual_rows)),
        "avatar_source_types": dict(Counter(row["visual"].get("avatar_source_type") for row in approved_visual_rows)),
        "existing_missing_live_visuals": [
            {
                "batch": row["batch"],
                "name": row["name"],
                "entity_id": row["canonical_entity_id"],
                "live_cover_count": row["visual"].get("live_cover_count"),
                "live_avatar_count": row["visual"].get("live_avatar_count"),
                "local_cover_exists": row["visual"]["cover"]["exists"],
                "local_avatar_exists": row["visual"]["avatar"]["exists"],
            }
            for row in existing_missing_live_visual
        ],
    }

    summary = {
        "space_id": SPACE_ID,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "expected_agents": len(agents),
        "by_batch": dict(Counter(agent["batch"] for agent in agents)),
        "enrich_existing": enrich_existing_count,
        "create_agent": missing_create_count,
        "duplicate_name_groups": len(duplicate_cleanup),
        "visual_issues": len(visual_issues),
        "missing_taxonomy_visual_files": len(missing_taxonomy_visual_files),
        "rows_needing_repositories": sum(1 for row in rows if row["relation_plan"].get("needs_repository")),
        "rows_needing_use_cases": sum(1 for row in rows if row["relation_plan"].get("needs_use_cases")),
        "rows_needing_provided_by": sum(1 for row in rows if row["relation_plan"].get("needs_provided_by")),
        "rows_needing_blocks": sum(1 for row in rows if row["relation_plan"].get("blocks_missing")),
    }

    report = {
        "summary": summary,
        "existing_only_summary": existing_only_summary,
        "target_specs": target_specs,
        "duplicate_cleanup": duplicate_cleanup,
        "visual_diff_summary": visual_diff_summary,
        "visual_issues": visual_issues,
        "missing_taxonomy_visual_files": missing_taxonomy_visual_files,
        "rows": rows,
    }
    write_json(OUT_JSON, report)

    lines = [
        "# ATG clean AI consolidated patch plan",
        "",
        f"Date: {DATE}",
        f"Space: `{SPACE_ID}`",
        "",
        "## Summary",
        "",
        f"- Expected agents: `{summary['expected_agents']}`",
        f"- By batch: `{summary['by_batch']}`",
        f"- Enrich existing canonical cards: `{summary['enrich_existing']}`",
        f"- Create missing agent cards: `{summary['create_agent']}`",
        f"- Duplicate groups needing cleanup: `{summary['duplicate_name_groups']}`",
        f"- Visual issues in approved/local manifests: `{summary['visual_issues']}`",
        f"- Missing taxonomy visual files: `{summary['missing_taxonomy_visual_files']}`",
        f"- Rows needing `Repositories`: `{summary['rows_needing_repositories']}`",
        f"- Rows needing `Provided by`: `{summary['rows_needing_provided_by']}`",
        f"- Rows needing `Use cases`: `{summary['rows_needing_use_cases']}`",
        f"- Rows needing comparison blocks: `{summary['rows_needing_blocks']}`",
        "",
        "## Next proposal scope: existing-only",
        "",
        f"- Existing batch rows: `{existing_only_summary['batch_rows']}`",
        f"- Unique canonical cards to enrich: `{existing_only_summary['unique_canonical_cards']}`",
        f"- Create rows intentionally excluded now: `{existing_only_summary['create_rows_excluded_from_next_proposal']}`",
        f"- Existing cards needing `Repositories`: `{existing_only_summary['rows_needing_repositories']}`",
        f"- Existing cards needing `Provided by`: `{existing_only_summary['rows_needing_provided_by']}`",
        f"- Existing cards needing `Use cases`: `{existing_only_summary['rows_needing_use_cases']}`",
        f"- Existing cards needing comparison blocks: `{existing_only_summary['rows_needing_blocks']}`",
        f"- Existing cards needing type normalization to `Agent + Project + Tool`: `{existing_only_summary['rows_needing_type_normalization']}`",
        f"- Existing-card local visual issues: `{existing_only_summary['local_visual_issues']}`",
        "",
        "Create-agent rows are useful for the final full ATG publication, but the next proposal should not create them.",
        "",
        "## Duplicate cleanup plan",
        "",
    ]
    if duplicate_cleanup:
        for item in duplicate_cleanup:
            lines.append(f"- `{item['name']}` canonical `{item['canonical']}`; cleanup Agent type on `{', '.join(item['cleanup_candidates'])}`")
    else:
        lines.append("- None")

    lines.extend(["", "## Visual issues", ""])
    if visual_issues:
        for item in visual_issues[:80]:
            lines.append(
                f"- `{item['name']}` / `{item['agent_id']}` approved={item['approved']} "
                f"coverExists={item['cover']['exists']} avatarExists={item['avatar']['exists']} "
                f"coverSource={item.get('cover_source_type')} avatarSource={item.get('avatar_source_type')}"
            )
        if len(visual_issues) > 80:
            lines.append(f"- ... {len(visual_issues) - 80} more in JSON report")
    else:
        lines.append("- None")

    lines.extend(["", "## Visual diff summary", ""])
    lines.extend(
        [
            f"- Approved agent visual rows: `{visual_diff_summary['approved_rows']}`",
            f"- Local approved cover+avatar complete: `{visual_diff_summary['local_visual_complete']}`",
            f"- Local approved cover/avatar missing: `{visual_diff_summary['local_visual_missing']}`",
            f"- Existing live cards with approved visuals: `{visual_diff_summary['existing_cards_with_approved_visuals']}`",
            f"- Existing live cards missing live cover or avatar: `{visual_diff_summary['existing_cards_missing_live_cover_or_avatar']}`",
            f"- New/create rows with approved complete visuals: `{visual_diff_summary['create_cards_with_approved_complete_visuals']}`",
            f"- Cover source types: `{visual_diff_summary['cover_source_types']}`",
            f"- Avatar source types: `{visual_diff_summary['avatar_source_types']}`",
        ]
    )
    if visual_diff_summary["existing_missing_live_visuals"]:
        lines.extend(["", "Existing cards missing live cover/avatar:"])
        for item in visual_diff_summary["existing_missing_live_visuals"]:
            lines.append(
                f"- `{item['name']}` `{item['entity_id']}` liveCover={item['live_cover_count']} "
                f"liveAvatar={item['live_avatar_count']} localCover={item['local_cover_exists']} "
                f"localAvatar={item['local_avatar_exists']}"
            )
    else:
        lines.append("- Existing cards missing live cover/avatar: none")

    lines.extend(["", "## Missing taxonomy visual files", ""])
    if missing_taxonomy_visual_files:
        for item in missing_taxonomy_visual_files[:80]:
            lines.append(f"- `{item['kind']}:{item['slug']}` cover={item['cover_exists']} avatar={item['avatar_exists']}")
        if len(missing_taxonomy_visual_files) > 80:
            lines.append(f"- ... {len(missing_taxonomy_visual_files) - 80} more in JSON report")
    else:
        lines.append("- None")

    lines.extend(["", "## Batch/action overview", "", "| Batch | Agent | Action | Canonical | Visual | Blocks missing |", "|---|---|---|---|---|---|"])
    for row in rows:
        visual_status = "ok" if row["visual"]["approved"] and row["visual"]["cover"]["exists"] and row["visual"]["avatar"]["exists"] else "needs_review"
        blocks = ", ".join(row["relation_plan"].get("blocks_missing", []))
        lines.append(
            f"| {row['batch']} | `{row['name']}` | {row['action']} | `{row['canonical_entity_id'] or ''}` | {visual_status} | {blocks} |"
        )

    lines.extend(
        [
            "",
            "## Notes",
            "",
            "- This is a read-only planner. It does not create Geo ops and does not publish.",
            "- Canonical duplicate selection uses the oldest `createdAt` among exact-name Agent-typed candidates.",
            "- The next live publisher should consume this plan, but still emit a dry-run ops file before any proposal.",
        ]
    )
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")

    print(json.dumps({"out_json": str(OUT_JSON), "out_md": str(OUT_MD), **summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
